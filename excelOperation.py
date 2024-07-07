from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os

import config

def columnFit(path):
    wb=load_workbook(path)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        lks=[]
        for i in range(1,ws.max_column+1):
            lk=1
            for j in range(1,ws.max_row+1):
                sz=ws.cell(row=j,column=i).value
                if isinstance(sz, (int, float)):
                    lk1=len(str(format(sz,',')))
                elif sz is None:
                    lk1=0
                else:
                    lk1=len(str(sz).encode('utf8'))
                if lk<lk1:
                    lk=lk1
            lks.append(lk)
        for i in range(1,ws.max_column+1):
            k=get_column_letter(i)
            ws.column_dimensions[k].width=min(lks[i-1],20)+2
    wb.save(path)
    wb.close()

def saveToExcel(dfs):
    if not os.path.exists("./achievements"):
        current_directory = os.path.dirname(__file__)
        imgs_path = os.path.join(current_directory, "achievements/")
        os.makedirs(imgs_path)
    writer = pd.ExcelWriter('achievements/achievements_'+config.language+'.xlsx', engine='xlsxwriter')

    for key in dfs:
        data = dfs[key]
        data['pair'].to_excel(writer, sheet_name=key, startrow=0, index=False)
        df = pd.DataFrame(data['notPair'], columns=['未匹配成就'])
        df.to_excel(writer, sheet_name=key, startrow=len(data['pair']) + 2, index=False)

    writer._save()

def readExcel(path=config.ACHIEVEMENTS_FILE):
    # 读取 Excel 文件
    xlsx_file = pd.ExcelFile(path)
    dfs = {}

    for sheet_name in xlsx_file.sheet_names:
        dfs[sheet_name] = xlsx_file.parse(sheet_name)

    return dfs